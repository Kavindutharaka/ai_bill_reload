import requests, json, os, base64
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

API_URL = "https://mas.phvtech.com/api/Master/sp"
LKR_OFFSET = timedelta(hours=5, minutes=30)
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "report_output")
IMAGES_DIR = os.path.join(OUTPUT_DIR, "bill_images_2026-03-27")
os.makedirs(IMAGES_DIR, exist_ok=True)

def query(sql):
    r = requests.post(API_URL, json={"SysID": sql})
    return r.json()

def to_lkr(dt_str):
    if not dt_str: return ""
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z",""))
        lkr = dt + LKR_OFFSET
        return lkr.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return dt_str

# Filter for March 27 LKR only (UTC: March 26 18:30 to March 27 18:30)
print("Fetching records for 2026-03-27 (LKR)...")
records = query("SELECT id, name, phone, bill_number, bill_date, shop_name, bill_total, eligible_total, is_eligible, items_json, capture_mode, reload_sent, created_at FROM rek_cc_reg WHERE DATEADD(MINUTE, 330, created_at) >= '2026-03-27 00:00:00' AND DATEADD(MINUTE, 330, created_at) < '2026-03-28 00:00:00' ORDER BY created_at DESC")
print(f"Found {len(records)} records for March 27")

# Styles
hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
hdr_fill = PatternFill("solid", fgColor="333333")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
green_fill = PatternFill("solid", fgColor="E8F8EF")
red_fill = PatternFill("solid", fgColor="FDECEA")
brand_fill = PatternFill("solid", fgColor="D4EDDA")
border = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD")
)
title_font = Font(bold=True, size=14, name="Arial", color="333333")
sub_font = Font(size=10, name="Arial", color="888888")

wb = Workbook()

# ===== SHEET 1: Summary =====
ws = wb.active
ws.title = "Summary"
ws.sheet_properties.tabColor = "FEC90D"
ws["A1"] = "Avurudu Reload Wasi - Daily Report (2026-03-27)"
ws["A1"].font = title_font
ws.merge_cells("A1:F1")
ws["A2"] = "Report Date: 2026-03-27 (LKR)"
ws["A2"].font = sub_font
ws.merge_cells("A2:F2")

total = len(records)
eligible = sum(1 for r in records if r.get("is_eligible"))
not_eligible = total - eligible
sent = sum(1 for r in records if r.get("reload_sent"))
pending = eligible - sent

summary_data = [
    ("Total Submissions", total),
    ("Eligible (Won)", eligible),
    ("Not Eligible (Lost)", not_eligible),
    ("Reloads Sent", sent),
    ("Reloads Pending", pending),
]

ws["A4"] = "METRIC"
ws["B4"] = "COUNT"
for c in ["A", "B"]:
    cell = ws[f"{c}4"]
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

for i, (label, val) in enumerate(summary_data, 5):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = Font(bold=True, size=11, name="Arial")
    ws[f"A{i}"].border = border
    ws[f"B{i}"] = val
    ws[f"B{i}"].font = Font(size=11, name="Arial")
    ws[f"B{i}"].alignment = Alignment(horizontal="center")
    ws[f"B{i}"].border = border

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 15

# ===== SHEET 2: All Submissions =====
ws2 = wb.create_sheet("All Submissions")
ws2.sheet_properties.tabColor = "3498DB"
headers = ["#", "Name", "Phone", "Date/Time (LKR)", "Bill Number", "Shop", "Bill Total", "Eligible Total", "Status", "Reload", "Capture Mode", "Image Folder"]
for col, h in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

for idx, r in enumerate(records, 1):
    row = idx + 1
    is_elig = r.get("is_eligible", False)
    fill = green_fill if is_elig else red_fill
    img_folder = f"bill_images_2026-03-27/record_{r['id']}"
    data = [
        idx, r.get("name", ""), r.get("phone", ""),
        to_lkr(r.get("created_at", "")), r.get("bill_number", ""),
        r.get("shop_name", "") or "-", r.get("bill_total", 0),
        r.get("eligible_total", 0),
        "ELIGIBLE" if is_elig else "NOT ELIGIBLE",
        "SENT" if r.get("reload_sent") else ("PENDING" if is_elig else "N/A"),
        (r.get("capture_mode") or "single").upper(), img_folder
    ]
    for col, val in enumerate(data, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = border
        cell.font = Font(size=10, name="Arial")
        if col in [7, 8]:
            cell.number_format = '#,##0.00'
        if col == 9:
            cell.font = Font(bold=True, size=10, name="Arial", color="155724" if is_elig else "721C24")

col_widths = [5, 20, 14, 22, 15, 18, 14, 14, 15, 12, 14, 22]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.auto_filter.ref = f"A1:L{len(records)+1}"
ws2.freeze_panes = "A2"

# ===== SHEET 3: Eligible Only =====
ws3 = wb.create_sheet("Eligible (Reload)")
ws3.sheet_properties.tabColor = "27AE60"
elig_headers = ["#", "Name", "Phone", "Date/Time (LKR)", "Bill Number", "Shop", "Eligible Total", "Reload Status", "Eligible Items"]
for col, h in enumerate(elig_headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

elig_idx = 0
for r in records:
    if not r.get("is_eligible"): continue
    elig_idx += 1
    row = elig_idx + 1
    items = []
    try: items = json.loads(r.get("items_json", "[]"))
    except: pass
    elig_items = [i for i in items if i.get("brand_matched") and i["brand_matched"] != "null"]
    elig_str = ", ".join([f"{i['item_name']} ({i['brand_matched']}) Rs {i.get('total_price',0)}" for i in elig_items])
    data = [
        elig_idx, r.get("name", ""), r.get("phone", ""),
        to_lkr(r.get("created_at", "")), r.get("bill_number", ""),
        r.get("shop_name", "") or "-", r.get("eligible_total", 0),
        "SENT" if r.get("reload_sent") else "PENDING", elig_str
    ]
    for col, val in enumerate(data, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.border = border
        cell.font = Font(size=10, name="Arial")
        if col == 7: cell.number_format = '#,##0.00'
        if col == 8:
            cell.font = Font(bold=True, size=10, name="Arial", color="155724" if r.get("reload_sent") else "856404")

elig_widths = [5, 20, 14, 22, 15, 18, 14, 14, 60]
for i, w in enumerate(elig_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.auto_filter.ref = f"A1:I{elig_idx+1}"
ws3.freeze_panes = "A2"

# ===== SHEET 4: Detailed Items =====
ws4 = wb.create_sheet("All Items Detail")
ws4.sheet_properties.tabColor = "E67E22"
item_headers = ["Record #", "Name", "Phone", "Bill #", "Item Name", "Brand Matched", "Qty", "Unit Price", "Total Price"]
for col, h in enumerate(item_headers, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border

item_row = 1
for r in records:
    items = []
    try: items = json.loads(r.get("items_json", "[]"))
    except: pass
    for item in items:
        item_row += 1
        is_brand = item.get("brand_matched") and item["brand_matched"] != "null"
        data = [
            r.get("id"), r.get("name", ""), r.get("phone", ""),
            r.get("bill_number", ""), item.get("item_name", ""),
            item.get("brand_matched", "") or "-", item.get("quantity", 0),
            item.get("unit_price", 0), item.get("total_price", 0)
        ]
        for col, val in enumerate(data, 1):
            cell = ws4.cell(row=item_row, column=col, value=val)
            cell.border = border
            cell.font = Font(size=10, name="Arial")
            if col in [8, 9]: cell.number_format = '#,##0.00'
            if is_brand:
                cell.fill = brand_fill
                if col == 6: cell.font = Font(bold=True, size=10, name="Arial", color="155724")

item_widths = [10, 18, 14, 14, 35, 14, 8, 12, 12]
for i, w in enumerate(item_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.auto_filter.ref = f"A1:I{item_row}"
ws4.freeze_panes = "A2"

# Save Excel
excel_path = os.path.join(OUTPUT_DIR, "Avurudu_Reload_Report_2026-03-27.xlsx")
wb.save(excel_path)
print(f"Excel saved: {excel_path}")

# ===== EXPORT BILL IMAGES =====
print("\nExporting bill images...")
for r in records:
    rec_id = r["id"]
    rec_dir = os.path.join(IMAGES_DIR, f"record_{rec_id}")
    os.makedirs(rec_dir, exist_ok=True)
    img_data = query(f"SELECT bill_image, bill_image_2, bill_image_3, capture_mode FROM rek_cc_reg WHERE id = {rec_id}")
    if not img_data: continue
    img = img_data[0]
    for i, col in enumerate(["bill_image", "bill_image_2", "bill_image_3"], 1):
        val = img.get(col, "")
        if val and val.startswith("data:image"):
            try:
                b64 = val.split(",", 1)[1]
                ext = "png" if "png" in val[:30] else "jpg"
                fname = f"bill_photo_{i}.{ext}"
                with open(os.path.join(rec_dir, fname), "wb") as f:
                    f.write(base64.b64decode(b64))
                print(f"  Record {rec_id}: saved {fname}")
            except Exception as e:
                print(f"  Record {rec_id}: error saving image {i}: {e}")
    with open(os.path.join(rec_dir, "info.txt"), "w", encoding="utf-8") as f:
        f.write(f"Name: {r.get('name','')}\n")
        f.write(f"Phone: {r.get('phone','')}\n")
        f.write(f"Bill #: {r.get('bill_number','')}\n")
        f.write(f"Shop: {r.get('shop_name','')}\n")
        f.write(f"Date (LKR): {to_lkr(r.get('created_at',''))}\n")
        f.write(f"Bill Total: Rs {r.get('bill_total',0)}\n")
        f.write(f"Eligible Total: Rs {r.get('eligible_total',0)}\n")
        f.write(f"Status: {'ELIGIBLE' if r.get('is_eligible') else 'NOT ELIGIBLE'}\n")
        f.write(f"Reload: {'SENT' if r.get('reload_sent') else 'PENDING' if r.get('is_eligible') else 'N/A'}\n")
        f.write(f"Capture Mode: {r.get('capture_mode','single')}\n")

print(f"\nDone! Excel: {excel_path}")
print(f"Images: {IMAGES_DIR}")
